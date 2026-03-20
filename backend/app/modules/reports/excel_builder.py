"""Excel report base class + child classes (openpyxl).

Base class: ExcelReport
  Methods: add_header(), add_table(), add_totals_row(),
  set_column_widths(), apply_conditional_formatting(), save_to_bytes()

Child classes — one per Excel report type from Раздел 7.
"""

from __future__ import annotations

import io
from abc import ABC, abstractmethod
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ═══════════════════════════════════════════════════════════════
# Base class
# ═══════════════════════════════════════════════════════════════

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_TOTAL_FONT = Font(bold=True, size=11)
_TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_NUM_FMT = "#,##0.00"


class ExcelReport(ABC):
    """Reusable base for every openpyxl-based report."""

    def __init__(self, title: str) -> None:
        self.wb = Workbook()
        self.ws: Worksheet = self.wb.active  # type: ignore[assignment]
        self.ws.title = title[:31]  # Excel tab-name limit
        self._current_row = 1

    # ── helpers ──────────────────────────────────────────────

    def add_header(self, title: str, subtitle: str | None = None) -> None:
        """Write report title (+ optional subtitle) at the top."""
        self.ws.merge_cells(
            start_row=self._current_row,
            start_column=1,
            end_row=self._current_row,
            end_column=8,
        )
        cell = self.ws.cell(row=self._current_row, column=1, value=title)
        cell.font = _TITLE_FONT
        self._current_row += 1

        if subtitle:
            self.ws.merge_cells(
                start_row=self._current_row,
                start_column=1,
                end_row=self._current_row,
                end_column=8,
            )
            self.ws.cell(row=self._current_row, column=1, value=subtitle)
            self._current_row += 1

        self._current_row += 1  # blank row

    def add_table(
        self,
        data: Sequence[dict[str, Any]],
        columns: list[tuple[str, str]],
        *,
        freeze_header: bool = True,
        number_columns: set[str] | None = None,
    ) -> int:
        """Write header row + data rows. Returns first data row number.

        Args:
            data: list of row-dicts
            columns: [(key, display_title), ...]
            freeze_header: pin header row
            number_columns: set of keys that need number formatting
        """
        number_columns = number_columns or set()
        header_row = self._current_row

        # Header
        for col_idx, (_, col_title) in enumerate(columns, start=1):
            cell = self.ws.cell(row=header_row, column=col_idx, value=col_title)
            cell.font = _HEADER_FONT_WHITE
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = _BORDER

        if freeze_header:
            self.ws.freeze_panes = f"A{header_row + 1}"

        self._current_row += 1
        first_data_row = self._current_row

        # Data
        for row_dict in data:
            for col_idx, (key, _) in enumerate(columns, start=1):
                value = row_dict.get(key)
                if isinstance(value, Decimal):
                    value = float(value)
                cell = self.ws.cell(row=self._current_row, column=col_idx, value=value)
                cell.border = _BORDER
                if key in number_columns and isinstance(value, (int, float)):
                    cell.number_format = _NUM_FMT
                    cell.alignment = Alignment(horizontal="right")
            self._current_row += 1

        return first_data_row

    def add_totals_row(
        self,
        totals: dict[str, Any],
        columns: list[tuple[str, str]],
        *,
        number_columns: set[str] | None = None,
    ) -> None:
        """Write bold totals row."""
        number_columns = number_columns or set()
        for col_idx, (key, _) in enumerate(columns, start=1):
            value = totals.get(key)
            if isinstance(value, Decimal):
                value = float(value)
            cell = self.ws.cell(row=self._current_row, column=col_idx, value=value)
            cell.font = _TOTAL_FONT
            cell.fill = _TOTAL_FILL
            cell.border = _BORDER
            if key in number_columns and isinstance(value, (int, float)):
                cell.number_format = _NUM_FMT
        self._current_row += 1

    def set_column_widths(self, widths: list[int]) -> None:
        for i, w in enumerate(widths, start=1):
            self.ws.column_dimensions[get_column_letter(i)].width = w

    def apply_conditional_formatting(
        self,
        col_letter: str,
        first_row: int,
        last_row: int,
        rules: list[tuple[str, str, Any, PatternFill]],
    ) -> None:
        """Apply conditional formatting rules.

        rules: [(operator, formula_or_value, _, fill), ...]
        operator: 'lessThan', 'greaterThan', 'equal', 'between'
        """
        rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
        for operator, value, _extra, fill in rules:
            self.ws.conditional_formatting.add(
                rng,
                CellIsRule(operator=operator, formula=[str(value)], fill=fill),
            )

    def save_to_bytes(self) -> bytes:
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ── abstract ──────────────────────────────────────────────

    @abstractmethod
    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        """Generate report bytes. Accepts AsyncSession + params dict."""


# ═══════════════════════════════════════════════════════════════
# Helper: execute text() query
# ═══════════════════════════════════════════════════════════════

async def _fetch(session: Any, sql: str, params: dict[str, Any]) -> list[dict]:
    from sqlalchemy import text

    result = await session.execute(text(sql), params)
    return [dict(row) for row in result.mappings().all()]


# ═══════════════════════════════════════════════════════════════
# ГРУППА 1: ФИНАНСОВЫЕ ОТЧЁТЫ
# ═══════════════════════════════════════════════════════════════

class BudgetBalanceReport(ExcelReport):
    """BUDGET_BALANCE — Остатки сметного лимита."""

    def __init__(self) -> None:
        super().__init__("Остатки лимита")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])
        level = params.get("level", "LS")
        date_to = params.get("date_to", date.today().isoformat())

        rows = await _fetch(session, """
            WITH item_facts AS (
                SELECT
                    ei.id AS item_id, ei.ls_id, ei.code, ei.name,
                    ei.amount_approved,
                    COALESCE(SUM(p.amount_fact), 0) AS fact_total,
                    COALESCE(SUM(CASE
                        WHEN make_date(p.year, p.month, 1) <= :date_to::date
                        THEN p.amount_fact ELSE 0
                    END), 0) AS fact_period
                FROM estimate_items ei
                LEFT JOIN progress p ON p.estimate_item_id = ei.id
                JOIN local_estimates le ON le.id = ei.ls_id
                JOIN local_estimate_bases leb ON leb.id = le.lsr_id
                JOIN object_estimates oe ON oe.id = leb.object_estimate_id
                JOIN summary_estimates se ON se.id = oe.summary_estimate_id
                WHERE se.project_id = :project_id
                GROUP BY ei.id, ei.ls_id, ei.code, ei.name, ei.amount_approved
            )
            SELECT
                code AS "Шифр", name AS "Наименование",
                amount_approved AS "Лимит",
                fact_total AS "Факт с начала",
                fact_period AS "Факт за период",
                (amount_approved - fact_total) AS "Остаток",
                CASE WHEN amount_approved > 0
                     THEN ROUND(fact_total * 100.0 / amount_approved, 2) ELSE 0
                END AS "percent"
            FROM item_facts
            ORDER BY code
        """, {"project_id": project_id, "date_to": date_to})

        columns = [
            ("Шифр", "Шифр"), ("Наименование", "Наименование"),
            ("Лимит", "Лимит"), ("Факт с начала", "Факт с начала"),
            ("Факт за период", "Факт за период"), ("Остаток", "Остаток"),
            ("percent", "% выполнения"),
        ]
        num_cols = {"Лимит", "Факт с начала", "Факт за период", "Остаток", "percent"}

        self.add_header("Остатки сметного лимита", f"Дата: {date_to}")
        first = self.add_table(rows, columns, number_columns=num_cols)

        if rows:
            last = first + len(rows) - 1
            self.apply_conditional_formatting("F", first, last, [
                ("lessThan", "0", None, _RED_FILL),
            ])

        self.set_column_widths([15, 40, 18, 18, 18, 18, 12])
        return self.save_to_bytes()


class BudgetByContractsReport(ExcelReport):
    """BUDGET_BY_CONTRACTS — Остатки лимитов в разрезе договоров."""

    def __init__(self) -> None:
        super().__init__("Лимиты по договорам")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])

        rows = await _fetch(session, """
            SELECT
                le.code AS "ЛС",
                ic.number AS "Договор",
                ic.counterparty AS "Контрагент",
                lic.percentage AS "Распределено %",
                ROUND(le.total_amount * lic.percentage / 100, 2) AS "Сумма по договору",
                COALESCE(fact.amount_fact, 0) AS "Факт",
                ROUND(le.total_amount * lic.percentage / 100, 2) -
                    COALESCE(fact.amount_fact, 0) AS "Остаток"
            FROM local_estimates le
            JOIN ls_income_contract lic ON lic.ls_id = le.id
            JOIN income_contracts ic ON ic.id = lic.income_contract_id
            JOIN local_estimate_bases leb ON leb.id = le.lsr_id
            JOIN object_estimates oe ON oe.id = leb.object_estimate_id
            JOIN summary_estimates se ON se.id = oe.summary_estimate_id
            LEFT JOIN LATERAL (
                SELECT SUM(p.amount_fact) AS amount_fact
                FROM progress p
                JOIN estimate_items ei ON ei.id = p.estimate_item_id
                WHERE ei.ls_id = le.id
            ) fact ON TRUE
            WHERE se.project_id = :project_id
            ORDER BY le.code, ic.number
        """, {"project_id": project_id})

        columns = [
            ("ЛС", "ЛС"), ("Договор", "Договор"), ("Контрагент", "Контрагент"),
            ("Распределено %", "% распр."),
            ("Сумма по договору", "Сумма по дог."), ("Факт", "Факт"),
            ("Остаток", "Остаток"),
        ]
        num_cols = {"Распределено %", "Сумма по договору", "Факт", "Остаток"}

        self.add_header("Остатки лимитов в разрезе договоров")
        self.add_table(rows, columns, number_columns=num_cols)
        self.set_column_widths([15, 18, 30, 12, 18, 18, 18])
        return self.save_to_bytes()


class MonthlyProgressReport(ExcelReport):
    """MONTHLY_PROGRESS — Помесячное выполнение (сводная)."""

    def __init__(self) -> None:
        super().__init__("Помесячное")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])
        year = params["year"]
        month_from = params.get("month_from", 1)
        month_to = params.get("month_to", 12)

        rows = await _fetch(session, """
            SELECT
                ei.code, ei.name, ei.amount_approved,
                p.month, COALESCE(SUM(p.amount_fact), 0) AS mfact
            FROM estimate_items ei
            LEFT JOIN progress p ON p.estimate_item_id = ei.id
                AND p.year = :year AND p.month BETWEEN :m1 AND :m2
            JOIN local_estimates le ON le.id = ei.ls_id
            JOIN local_estimate_bases leb ON leb.id = le.lsr_id
            JOIN object_estimates oe ON oe.id = leb.object_estimate_id
            JOIN summary_estimates se ON se.id = oe.summary_estimate_id
            WHERE se.project_id = :project_id
            GROUP BY ei.id, ei.code, ei.name, ei.amount_approved, p.month
            ORDER BY ei.code, p.month
        """, {"project_id": project_id, "year": year, "m1": month_from, "m2": month_to})

        # Pivot by month
        months = list(range(month_from, month_to + 1))
        month_names = [
            "", "Янв", "Фев", "Мар", "Апр", "Май", "Июн",
            "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек",
        ]

        pivoted: dict[str, dict[str, Any]] = {}
        for r in rows:
            key = r["code"]
            if key not in pivoted:
                pivoted[key] = {
                    "code": r["code"], "name": r["name"],
                    "limit": r["amount_approved"],
                }
                for m in months:
                    pivoted[key][f"m{m}"] = Decimal(0)
                pivoted[key]["total"] = Decimal(0)
            if r["month"] is not None:
                pivoted[key][f"m{r['month']}"] = r["mfact"]
                pivoted[key]["total"] = pivoted[key].get("total", Decimal(0)) + r["mfact"]

        data = list(pivoted.values())
        for d in data:
            d["remain"] = d["limit"] - d["total"]

        columns: list[tuple[str, str]] = [("code", "Статья"), ("name", "Наименование"), ("limit", "Лимит")]
        for m in months:
            columns.append((f"m{m}", month_names[m]))
        columns.extend([("total", "Итого"), ("remain", "Остаток")])

        num_cols = {"limit", "total", "remain"} | {f"m{m}" for m in months}

        self.add_header(f"Помесячное выполнение за {year} г.",
                        f"Месяцы: {month_from}–{month_to}")
        self.add_table(data, columns, number_columns=num_cols)
        widths = [15, 35, 16] + [12] * len(months) + [14, 14]
        self.set_column_widths(widths)
        return self.save_to_bytes()


class QuarterlyProgressReport(ExcelReport):
    """QUARTERLY_PROGRESS — Освоение нарастающим итогом (поквартально)."""

    def __init__(self) -> None:
        super().__init__("Поквартально")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])
        year = params["year"]

        rows = await _fetch(session, """
            SELECT
                ei.code, ei.name, ei.amount_approved,
                EXTRACT(QUARTER FROM make_date(p.year, p.month, 1))::int AS q,
                COALESCE(SUM(p.amount_fact), 0) AS qfact
            FROM estimate_items ei
            LEFT JOIN progress p ON p.estimate_item_id = ei.id AND p.year = :year
            JOIN local_estimates le ON le.id = ei.ls_id
            JOIN local_estimate_bases leb ON leb.id = le.lsr_id
            JOIN object_estimates oe ON oe.id = leb.object_estimate_id
            JOIN summary_estimates se ON se.id = oe.summary_estimate_id
            WHERE se.project_id = :project_id
            GROUP BY ei.id, ei.code, ei.name, ei.amount_approved, q
            ORDER BY ei.code, q
        """, {"project_id": project_id, "year": year})

        pivoted: dict[str, dict[str, Any]] = {}
        for r in rows:
            key = r["code"]
            if key not in pivoted:
                pivoted[key] = {
                    "code": r["code"], "name": r["name"],
                    "limit": r["amount_approved"],
                    "q1": Decimal(0), "q2": Decimal(0),
                    "q3": Decimal(0), "q4": Decimal(0), "total": Decimal(0),
                }
            if r["q"] is not None:
                pivoted[key][f"q{r['q']}"] = r["qfact"]
                pivoted[key]["total"] += r["qfact"]

        data = list(pivoted.values())
        for d in data:
            d["remain"] = d["limit"] - d["total"]

        columns = [
            ("code", "Статья"), ("name", "Наименование"), ("limit", "Лимит"),
            ("q1", "Q1"), ("q2", "Q2"), ("q3", "Q3"), ("q4", "Q4"),
            ("total", "Итого"), ("remain", "Остаток"),
        ]
        num_cols = {"limit", "q1", "q2", "q3", "q4", "total", "remain"}

        self.add_header(f"Освоение поквартально за {year} г.")
        self.add_table(data, columns, number_columns=num_cols)
        self.set_column_widths([15, 35, 16, 14, 14, 14, 14, 14, 14])
        return self.save_to_bytes()


# ═══════════════════════════════════════════════════════════════
# ГРУППА 2 (Excel): КС-6а
# ═══════════════════════════════════════════════════════════════

class KS6AReport(ExcelReport):
    """KS6A — Журнал учёта выполненных работ (КС-6а)."""

    def __init__(self) -> None:
        super().__init__("КС-6а")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])
        ls_id = str(params["ls_id"])

        rows = await _fetch(session, """
            SELECT
                ei.code, ei.name, ei.unit,
                ei.volume_approved, ei.amount_approved,
                p.year, p.month, p.volume_fact, p.amount_fact,
                SUM(p.volume_fact) OVER (
                    PARTITION BY ei.id ORDER BY p.year, p.month
                    ROWS UNBOUNDED PRECEDING
                ) AS cumulative_volume,
                SUM(p.amount_fact) OVER (
                    PARTITION BY ei.id ORDER BY p.year, p.month
                    ROWS UNBOUNDED PRECEDING
                ) AS cumulative_amount
            FROM estimate_items ei
            JOIN progress p ON p.estimate_item_id = ei.id
            WHERE ei.ls_id = :ls_id
            ORDER BY ei.code, p.year, p.month
        """, {"ls_id": ls_id})

        columns = [
            ("code", "Шифр"), ("name", "Наименование"), ("unit", "Ед."),
            ("volume_approved", "V утв."), ("amount_approved", "Сумма утв."),
            ("year", "Год"), ("month", "Мес."),
            ("volume_fact", "V факт"), ("amount_fact", "Сумма факт"),
            ("cumulative_volume", "V нараст."), ("cumulative_amount", "Сумма нараст."),
        ]
        num_cols = {
            "volume_approved", "amount_approved", "volume_fact",
            "amount_fact", "cumulative_volume", "cumulative_amount",
        }

        self.add_header("Журнал учёта выполненных работ (КС-6а)")
        self.add_table(rows, columns, number_columns=num_cols)
        self.set_column_widths([12, 35, 8, 12, 15, 8, 6, 12, 15, 12, 15])
        return self.save_to_bytes()


# ═══════════════════════════════════════════════════════════════
# ГРУППА 3: МАТЕРИАЛЬНАЯ ОТЧЁТНОСТЬ
# ═══════════════════════════════════════════════════════════════

class M29ExcelReport(ExcelReport):
    """M29_REPORT — Отчёт о расходе материалов (Excel)."""

    def __init__(self) -> None:
        super().__init__("Форма М-29")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        report_id = str(params.get("report_id") or "")
        project_id = str(params.get("project_id", ""))

        if report_id:
            where = "mr.id = :report_id"
            qp = {"report_id": report_id}
        else:
            where = "mr.project_id = :project_id AND mr.period_month = :month AND mr.period_year = :year"
            qp = {"project_id": project_id, "month": params["month"], "year": params["year"]}

        rows = await _fetch(session, f"""
            SELECT
                s.name AS material, s.mark_standard AS standard, ml.unit,
                ml.norm_consumption, ml.actual_consumption, ml.deviation,
                mr.foreman_name, mr.period_month, mr.period_year
            FROM m29_lines ml
            JOIN m29_reports mr ON mr.id = ml.report_id
            JOIN specifications s ON s.id = ml.specification_id
            WHERE {where}
            ORDER BY s.position
        """, qp)

        columns = [
            ("material", "Материал"), ("standard", "ГОСТ/ТУ"), ("unit", "Ед."),
            ("norm_consumption", "Норма"), ("actual_consumption", "Факт"),
            ("deviation", "Отклонение"),
        ]
        num_cols = {"norm_consumption", "actual_consumption", "deviation"}

        period = ""
        if rows:
            period = f"{rows[0].get('period_month', '')}/{rows[0].get('period_year', '')}"
        self.add_header("Отчёт о расходе материалов (М-29)", f"Период: {period}")
        first = self.add_table(rows, columns, number_columns=num_cols)

        if rows:
            last = first + len(rows) - 1
            self.apply_conditional_formatting("F", first, last, [
                ("greaterThan", "0", None, _RED_FILL),
                ("equal", "0", None, _GREEN_FILL),
            ])

        self.set_column_widths([30, 20, 8, 14, 14, 14])
        return self.save_to_bytes()


class MaterialDemandReport(ExcelReport):
    """MATERIAL_DEMAND — Сводная потребность в материалах."""

    def __init__(self) -> None:
        super().__init__("Потребность")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])
        ls_filter = ""
        qp: dict[str, Any] = {"project_id": project_id}
        if params.get("ls_id"):
            ls_filter = "AND d.id IN (SELECT ld.drawing_id FROM ls_drawing ld WHERE ld.ls_id = :ls_id)"
            qp["ls_id"] = str(params["ls_id"])

        rows = await _fetch(session, f"""
            SELECT
                s.position, s.name, s.mark_standard AS standard,
                s.unit, s.volume_designed,
                COALESCE(cert.accepted, 0) AS accepted,
                s.volume_designed - COALESCE(cert.accepted, 0) AS deficit
            FROM specifications s
            JOIN drawings d ON d.id = s.drawing_id
            JOIN summary_estimates se ON se.project_id = :project_id
            LEFT JOIN LATERAL (
                SELECT COUNT(*) AS accepted
                FROM material_certificates mc
                WHERE mc.specification_id = s.id AND mc.status = 'ACCEPTED'
            ) cert ON TRUE
            WHERE d.project_id = :project_id {ls_filter}
            ORDER BY s.position
        """, qp)

        columns = [
            ("position", "Позиция"), ("name", "Наименование"),
            ("standard", "ГОСТ/ТУ"), ("unit", "Ед."),
            ("volume_designed", "По проекту"), ("accepted", "Поступило"),
            ("deficit", "Дефицит"),
        ]
        num_cols = {"volume_designed", "accepted", "deficit"}

        self.add_header("Сводная потребность в материалах")
        first = self.add_table(rows, columns, number_columns=num_cols)
        if rows:
            self.apply_conditional_formatting(
                "G", first, first + len(rows) - 1,
                [("greaterThan", "0", None, _YELLOW_FILL)],
            )
        self.set_column_widths([10, 35, 20, 8, 14, 14, 14])
        return self.save_to_bytes()


# ═══════════════════════════════════════════════════════════════
# ГРУППА 4: СТАТУСЫ ДОКУМЕНТАЦИИ
# ═══════════════════════════════════════════════════════════════

class IDRegistryReport(ExcelReport):
    """ID_REGISTRY — Реестр исполнительной документации."""

    def __init__(self) -> None:
        super().__init__("Реестр ИД")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])
        status_filter = ""
        qp: dict[str, Any] = {"project_id": project_id}
        if params.get("status"):
            status_filter = "AND ad.status = :status"
            qp["status"] = params["status"]

        rows = await _fetch(session, f"""
            SELECT
                ad.number, ad.type, ad.name,
                ad.work_date, ad.sign_date,
                ad.foreman, ad.supervisor, ad.status
            FROM asbuilt_docs ad
            WHERE ad.project_id = :project_id {status_filter}
            ORDER BY ad.number
        """, qp)

        columns = [
            ("number", "№"), ("type", "Тип"), ("name", "Наименование"),
            ("work_date", "Дата работ"), ("sign_date", "Дата подп."),
            ("foreman", "Прораб"), ("supervisor", "Технадзор"),
            ("status", "Статус"),
        ]

        self.add_header("Реестр исполнительной документации")
        self.add_table(rows, columns)
        self.set_column_widths([10, 10, 35, 14, 14, 20, 20, 14])
        return self.save_to_bytes()


class IDReadinessReport(ExcelReport):
    """ID_READINESS — Дашборд готовности пакета ИД к сдаче (3 листа)."""

    def __init__(self) -> None:
        super().__init__("Готовность ИД")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])

        # Sheet 1: сводная по разделам (марка чертежа)
        sheet1_rows = await _fetch(session, """
            SELECT
                d.mark,
                COUNT(DISTINCT d.id) AS drawings_count,
                COUNT(DISTINCT d.id) AS aosr_required,
                COUNT(DISTINCT CASE WHEN ad.status = 'SIGNED' THEN ad.id END) AS aosr_done,
                CASE WHEN COUNT(DISTINCT d.id) > 0
                     THEN ROUND(COUNT(DISTINCT CASE WHEN ad.status = 'SIGNED' THEN ad.id END) * 100.0
                          / COUNT(DISTINCT d.id), 1)
                     ELSE 0
                END AS readiness_pct
            FROM drawings d
            LEFT JOIN asbuilt_drawing abd ON abd.drawing_id = d.id
            LEFT JOIN asbuilt_docs ad ON ad.id = abd.asbuilt_doc_id
            WHERE d.project_id = :project_id
            GROUP BY d.mark
            ORDER BY d.mark
        """, {"project_id": project_id})

        columns1 = [
            ("mark", "Марка"), ("drawings_count", "Чертежей"),
            ("aosr_required", "АОСР треб."), ("aosr_done", "АОСР оформл."),
            ("readiness_pct", "% готовн."),
        ]
        self.add_header("Готовность пакета ИД — Сводная")
        self.add_table(sheet1_rows, columns1, number_columns={"readiness_pct"})
        self.set_column_widths([15, 14, 14, 14, 12])

        # Sheet 2: чертежи без АОСР
        ws2 = self.wb.create_sheet("Без АОСР")
        self._current_row = 1
        old_ws = self.ws
        self.ws = ws2

        sheet2_rows = await _fetch(session, """
            SELECT d.mark, d.name, d.sheet_number, d.status
            FROM drawings d
            WHERE d.project_id = :project_id
              AND NOT EXISTS (
                  SELECT 1 FROM asbuilt_drawing abd
                  JOIN asbuilt_docs ad ON ad.id = abd.asbuilt_doc_id
                  WHERE abd.drawing_id = d.id AND ad.status = 'SIGNED'
              )
            ORDER BY d.mark, d.sheet_number
        """, {"project_id": project_id})

        columns2 = [
            ("mark", "Марка"), ("name", "Наименование"),
            ("sheet_number", "Лист"), ("status", "Статус"),
        ]
        self.add_header("Чертежи без привязанных АОСР")
        self.add_table(sheet2_rows, columns2)
        self.set_column_widths([15, 40, 10, 14])

        # Sheet 3: статус сертификатов
        ws3 = self.wb.create_sheet("Сертификаты")
        self._current_row = 1
        self.ws = ws3

        sheet3_rows = await _fetch(session, """
            SELECT
                s.name AS material, s.mark_standard,
                mc.document_number, mc.issue_date, mc.expiry_date, mc.status
            FROM material_certificates mc
            JOIN specifications s ON s.id = mc.specification_id
            JOIN drawings d ON d.id = s.drawing_id
            WHERE d.project_id = :project_id
            ORDER BY s.name
        """, {"project_id": project_id})

        columns3 = [
            ("material", "Материал"), ("mark_standard", "ГОСТ/ТУ"),
            ("document_number", "№ серт."),
            ("issue_date", "Дата выдачи"), ("expiry_date", "Срок"),
            ("status", "Статус"),
        ]
        self.add_header("Статус сертификатов по спецификациям")
        self.add_table(sheet3_rows, columns3)
        self.set_column_widths([30, 20, 15, 14, 14, 14])

        self.ws = old_ws
        return self.save_to_bytes()


class InspectionExcelReport(ExcelReport):
    """INSPECTION_REPORT — Сводный отчёт по предписаниям СДО."""

    def __init__(self) -> None:
        super().__init__("Предписания")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])
        qp: dict[str, Any] = {"project_id": project_id}
        filters = ""
        if params.get("status"):
            filters += " AND i.status = :status"
            qp["status"] = params["status"]
        if params.get("date_from"):
            filters += " AND i.created_at >= :date_from::date"
            qp["date_from"] = params["date_from"]
        if params.get("date_to"):
            filters += " AND i.created_at <= :date_to::date"
            qp["date_to"] = params["date_to"]

        rows = await _fetch(session, f"""
            SELECT
                i.number, i.description, i.norm_reference,
                i.planned_fix_date, i.actual_fix_date,
                i.status,
                CASE WHEN i.status != 'CLOSED'
                     AND i.planned_fix_date < CURRENT_DATE
                     THEN 'Да' ELSE 'Нет'
                END AS overdue
            FROM inspections i
            WHERE i.project_id = :project_id {filters}
            ORDER BY i.number
        """, qp)

        columns = [
            ("number", "№"), ("description", "Описание"),
            ("norm_reference", "Норматив"),
            ("planned_fix_date", "Плановый срок"),
            ("actual_fix_date", "Факт. срок"),
            ("status", "Статус"), ("overdue", "Просрочено"),
        ]

        self.add_header("Сводный отчёт по предписаниям СДО")
        first = self.add_table(rows, columns)

        # Totals: counters by status
        from collections import Counter
        status_counts = Counter(r.get("status") for r in rows)
        overdue_count = sum(1 for r in rows if r.get("overdue") == "Да")
        totals = {
            "number": "ИТОГО:",
            "description": f"Всего: {len(rows)}",
            "norm_reference": f"Просрочено: {overdue_count}",
        }
        for st, cnt in sorted(status_counts.items()):
            totals["status"] = totals.get("status", "") + f"{st}: {cnt}  "
        self.add_totals_row(totals, columns)

        self.set_column_widths([8, 40, 25, 14, 14, 14, 12])
        return self.save_to_bytes()


class CertificateStatusReport(ExcelReport):
    """CERTIFICATE_STATUS — Реестр сертификатов и паспортов качества."""

    def __init__(self) -> None:
        super().__init__("Сертификаты")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])
        qp: dict[str, Any] = {"project_id": project_id}
        filt = ""
        if params.get("status"):
            filt = "AND mc.status = :status"
            qp["status"] = params["status"]

        rows = await _fetch(session, f"""
            SELECT
                s.name AS material, s.mark_standard,
                mc.document_number, mc.issue_date, mc.expiry_date,
                mc.status,
                CASE WHEN mc.expiry_date < CURRENT_DATE THEN 'Просрочен'
                     WHEN mc.expiry_date < CURRENT_DATE + INTERVAL '30 days' THEN '<30 дней'
                     ELSE ''
                END AS expiry_flag
            FROM material_certificates mc
            JOIN specifications s ON s.id = mc.specification_id
            JOIN drawings d ON d.id = s.drawing_id
            WHERE d.project_id = :project_id {filt}
            ORDER BY mc.expiry_date
        """, qp)

        columns = [
            ("material", "Материал"), ("mark_standard", "ГОСТ/ТУ"),
            ("document_number", "№ серт."), ("issue_date", "Дата"),
            ("expiry_date", "Срок"), ("status", "Статус"),
            ("expiry_flag", "Истекает"),
        ]

        self.add_header("Реестр сертификатов и паспортов качества")
        first = self.add_table(rows, columns)

        if rows:
            last = first + len(rows) - 1
            # Manual cell coloring for expiry_flag column (G)
            for i, r in enumerate(rows):
                flag = r.get("expiry_flag", "")
                row_num = first + i
                if flag == "Просрочен":
                    self.ws.cell(row=row_num, column=7).fill = _RED_FILL
                elif flag == "<30 дней":
                    self.ws.cell(row=row_num, column=7).fill = _YELLOW_FILL

        self.set_column_widths([30, 20, 15, 14, 14, 14, 12])
        return self.save_to_bytes()


# ═══════════════════════════════════════════════════════════════
# ГРУППА 5: СРОКИ И ПЛАНИРОВАНИЕ
# ═══════════════════════════════════════════════════════════════

class ScheduleDeviationReport(ExcelReport):
    """SCHEDULE_DEVIATION — Отчёт об отклонении от графика."""

    def __init__(self) -> None:
        super().__init__("Отклонения")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])

        rows = await _fetch(session, """
            SELECT
                le.code AS ls_code,
                p.planned_end,
                CASE WHEN le.total_amount > 0
                     THEN ROUND(COALESCE(fact.total_fact, 0) * 100.0 / le.total_amount, 2)
                     ELSE 0
                END AS fact_pct,
                CASE WHEN COALESCE(fact.total_fact, 0) > 0 AND le.total_amount > 0
                     THEN CURRENT_DATE + (
                         (p.planned_end - p.planned_start)
                         * (1 - COALESCE(fact.total_fact, 0)::numeric / le.total_amount)
                     )::int
                     ELSE p.planned_end
                END AS eac_date,
                CASE WHEN COALESCE(fact.total_fact, 0) > 0 AND le.total_amount > 0
                     THEN (
                         CURRENT_DATE + (
                             (p.planned_end - p.planned_start)
                             * (1 - COALESCE(fact.total_fact, 0)::numeric / le.total_amount)
                         )::int
                     ) - p.planned_end
                     ELSE 0
                END AS deviation_days,
                CASE
                    WHEN COALESCE(fact.total_fact, 0) > 0 AND le.total_amount > 0
                         AND (CURRENT_DATE + (
                             (p.planned_end - p.planned_start)
                             * (1 - COALESCE(fact.total_fact, 0)::numeric / le.total_amount)
                         )::int) - p.planned_end > 30
                    THEN 'Критично'
                    WHEN COALESCE(fact.total_fact, 0) > 0 AND le.total_amount > 0
                         AND (CURRENT_DATE + (
                             (p.planned_end - p.planned_start)
                             * (1 - COALESCE(fact.total_fact, 0)::numeric / le.total_amount)
                         )::int) - p.planned_end > 0
                    THEN 'Отставание'
                    ELSE 'В графике'
                END AS deviation_status
            FROM local_estimates le
            JOIN local_estimate_bases leb ON leb.id = le.lsr_id
            JOIN object_estimates oe ON oe.id = leb.object_estimate_id
            JOIN summary_estimates se ON se.id = oe.summary_estimate_id
            JOIN projects p ON p.id = se.project_id
            LEFT JOIN LATERAL (
                SELECT SUM(pr.amount_fact) AS total_fact
                FROM progress pr
                JOIN estimate_items ei ON ei.id = pr.estimate_item_id
                WHERE ei.ls_id = le.id
            ) fact ON TRUE
            WHERE se.project_id = :project_id
            ORDER BY le.code
        """, {"project_id": project_id})

        columns = [
            ("ls_code", "ЛС"), ("planned_end", "Плановый срок"),
            ("fact_pct", "Факт %"), ("eac_date", "Расч. завершение"),
            ("deviation_days", "Отставание (дней)"),
            ("deviation_status", "Статус"),
        ]
        num_cols = {"fact_pct", "deviation_days"}

        self.add_header("Отчёт об отклонении от графика")
        first = self.add_table(rows, columns, number_columns=num_cols)

        if rows:
            for i, r in enumerate(rows):
                st = r.get("deviation_status", "")
                row_num = first + i
                if st == "Критично":
                    self.ws.cell(row=row_num, column=6).fill = _RED_FILL
                elif st == "Отставание":
                    self.ws.cell(row=row_num, column=6).fill = _YELLOW_FILL
                else:
                    self.ws.cell(row=row_num, column=6).fill = _GREEN_FILL

        self.set_column_widths([15, 16, 12, 16, 16, 14])
        return self.save_to_bytes()


# ═══════════════════════════════════════════════════════════════
# ГРУППА 6: ЖУРНАЛЫ СДО (Excel)
# ═══════════════════════════════════════════════════════════════

class InputControlJournalReport(ExcelReport):
    """INPUT_CONTROL_JOURNAL — Журнал входного контроля."""

    def __init__(self) -> None:
        super().__init__("Входной контроль")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])
        qp: dict[str, Any] = {"project_id": project_id}
        filt = ""
        if params.get("date_from"):
            filt += " AND mc.issue_date >= :date_from::date"
            qp["date_from"] = params["date_from"]
        if params.get("date_to"):
            filt += " AND mc.issue_date <= :date_to::date"
            qp["date_to"] = params["date_to"]

        rows = await _fetch(session, f"""
            SELECT
                ROW_NUMBER() OVER (ORDER BY mc.issue_date) AS num,
                s.name AS material, s.mark_standard,
                mc.issue_date, mc.document_number,
                mc.status AS result,
                CASE WHEN mc.status = 'ACCEPTED' THEN 'Принят'
                     WHEN mc.status = 'REJECTED' THEN 'Отклонён: ' || COALESCE(mc.rejection_reason, '')
                     ELSE mc.status
                END AS decision
            FROM material_certificates mc
            JOIN specifications s ON s.id = mc.specification_id
            JOIN drawings d ON d.id = s.drawing_id
            WHERE d.project_id = :project_id {filt}
            ORDER BY mc.issue_date
        """, qp)

        columns = [
            ("num", "№"), ("material", "Материал"), ("mark_standard", "ГОСТ/ТУ"),
            ("issue_date", "Дата"), ("document_number", "№ серт."),
            ("result", "Результат"), ("decision", "Решение"),
        ]

        self.add_header("Журнал входного контроля")
        self.add_table(rows, columns)
        self.set_column_widths([6, 30, 20, 14, 15, 14, 30])
        return self.save_to_bytes()


# ═══════════════════════════════════════════════════════════════
# ГРУППА 7: ДОГОВОРНЫЕ ОТЧЁТЫ
# ═══════════════════════════════════════════════════════════════

class ContractRegistryReport(ExcelReport):
    """CONTRACT_REGISTRY — Реестр договоров проекта."""

    def __init__(self) -> None:
        super().__init__("Реестр договоров")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])
        contract_type = params.get("contract_type", "ALL")

        rows_income = []
        rows_expense = []

        if contract_type in ("INCOME", "ALL"):
            rows_income = await _fetch(session, """
                SELECT
                    ic.number, ic.counterparty, 'Доходный' AS type,
                    ic.total_amount, ic.status, ic.end_date,
                    COALESCE(fact.total_fact, 0) AS executed,
                    ic.total_amount - COALESCE(fact.total_fact, 0) AS remaining
                FROM income_contracts ic
                LEFT JOIN LATERAL (
                    SELECT SUM(p.amount_fact) AS total_fact
                    FROM progress p
                    JOIN estimate_items ei ON ei.id = p.estimate_item_id
                    JOIN local_estimates le ON le.id = ei.ls_id
                    JOIN ls_income_contract lic ON lic.ls_id = le.id
                    WHERE lic.income_contract_id = ic.id
                ) fact ON TRUE
                WHERE ic.project_id = :project_id
                ORDER BY ic.number
            """, {"project_id": project_id})

        if contract_type in ("EXPENSE", "ALL"):
            rows_expense = await _fetch(session, """
                SELECT
                    ec.number, ec.counterparty, 'Расходный' AS type,
                    ec.total_amount, ec.status, ec.end_date,
                    0 AS executed,
                    ec.total_amount AS remaining
                FROM expense_contracts ec
                WHERE ec.project_id = :project_id
                ORDER BY ec.number
            """, {"project_id": project_id})

        all_rows = rows_income + rows_expense

        columns = [
            ("number", "№ договора"), ("counterparty", "Контрагент"),
            ("type", "Тип"), ("total_amount", "Сумма"),
            ("executed", "Выполнено"), ("remaining", "Остаток"),
            ("status", "Статус"), ("end_date", "Срок оконч."),
        ]
        num_cols = {"total_amount", "executed", "remaining"}

        self.add_header("Реестр договоров проекта")
        self.add_table(all_rows, columns, number_columns=num_cols)
        self.set_column_widths([18, 30, 12, 18, 18, 18, 14, 14])
        return self.save_to_bytes()


class MarginReport(ExcelReport):
    """MARGIN_REPORT — Экономия/перерасход (маржа генподрядчика)."""

    def __init__(self) -> None:
        super().__init__("Маржа")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])

        rows = await _fetch(session, """
            SELECT
                le.code AS ls,
                ic.number AS income_contract,
                ec.number AS expense_contract,
                ROUND(le.total_amount * lic.percentage / 100, 2) AS income_amount,
                ROUND(ec.total_amount * iec.percentage / 100, 2) AS expense_amount,
                ROUND(le.total_amount * lic.percentage / 100, 2)
                    - ROUND(ec.total_amount * iec.percentage / 100, 2) AS margin_rub,
                CASE WHEN le.total_amount * lic.percentage > 0
                     THEN ROUND(
                         (le.total_amount * lic.percentage / 100
                          - ec.total_amount * iec.percentage / 100)
                         * 100.0 / (le.total_amount * lic.percentage / 100), 2
                     )
                     ELSE 0
                END AS margin_pct
            FROM local_estimates le
            JOIN ls_income_contract lic ON lic.ls_id = le.id
            JOIN income_contracts ic ON ic.id = lic.income_contract_id
            JOIN income_expense_contract iec ON iec.income_contract_id = ic.id
            JOIN expense_contracts ec ON ec.id = iec.expense_contract_id
            JOIN local_estimate_bases leb ON leb.id = le.lsr_id
            JOIN object_estimates oe ON oe.id = leb.object_estimate_id
            JOIN summary_estimates se ON se.id = oe.summary_estimate_id
            WHERE se.project_id = :project_id
            ORDER BY le.code
        """, {"project_id": project_id})

        columns = [
            ("ls", "ЛС"), ("income_contract", "Доходный дог."),
            ("expense_contract", "Расходный дог."),
            ("income_amount", "Доход (руб.)"), ("expense_amount", "Расход (руб.)"),
            ("margin_rub", "Маржа (руб.)"), ("margin_pct", "Маржа (%)"),
        ]
        num_cols = {"income_amount", "expense_amount", "margin_rub", "margin_pct"}

        self.add_header("Маржа генподрядчика")
        first = self.add_table(rows, columns, number_columns=num_cols)

        if rows:
            for i, r in enumerate(rows):
                val = r.get("margin_rub", 0) or 0
                row_num = first + i
                if val < 0:
                    self.ws.cell(row=row_num, column=6).fill = _RED_FILL
                elif val > 0:
                    self.ws.cell(row=row_num, column=6).fill = _GREEN_FILL

        self.set_column_widths([15, 18, 18, 16, 16, 16, 12])
        return self.save_to_bytes()


class UncoveredLSReport(ExcelReport):
    """UNCOVERED_LS — ЛС с незакрытыми объёмами (без договора)."""

    def __init__(self) -> None:
        super().__init__("Незакрытые ЛС")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        project_id = str(params["project_id"])

        rows = await _fetch(session, """
            SELECT
                le.code AS ls,
                le.total_amount,
                COALESCE(dist.total_pct, 0) AS distributed_pct,
                ROUND(le.total_amount * (100 - COALESCE(dist.total_pct, 0)) / 100, 2)
                    AS uncovered_amount
            FROM local_estimates le
            JOIN local_estimate_bases leb ON leb.id = le.lsr_id
            JOIN object_estimates oe ON oe.id = leb.object_estimate_id
            JOIN summary_estimates se ON se.id = oe.summary_estimate_id
            LEFT JOIN LATERAL (
                SELECT SUM(lic.percentage) AS total_pct
                FROM ls_income_contract lic
                WHERE lic.ls_id = le.id
            ) dist ON TRUE
            WHERE se.project_id = :project_id
              AND COALESCE(dist.total_pct, 0) < 100
            ORDER BY le.code
        """, {"project_id": project_id})

        columns = [
            ("ls", "ЛС"), ("total_amount", "Сумма ЛС"),
            ("distributed_pct", "Распределено (%)"),
            ("uncovered_amount", "Нераспр. остаток (руб.)"),
        ]
        num_cols = {"total_amount", "distributed_pct", "uncovered_amount"}

        self.add_header("ЛС с незакрытыми объёмами")
        self.add_table(rows, columns, number_columns=num_cols)
        self.set_column_widths([15, 18, 18, 22])
        return self.save_to_bytes()


# ═══════════════════════════════════════════════════════════════
# ГРУППА 8: УПРАВЛЕНЧЕСКИЕ (Excel)
# ═══════════════════════════════════════════════════════════════

class MultiProjectSummaryReport(ExcelReport):
    """MULTI_PROJECT_SUMMARY — Сводная по всем проектам."""

    def __init__(self) -> None:
        super().__init__("Сводная")

    async def build(self, session: Any, params: dict[str, Any]) -> bytes:
        rows = await _fetch(session, """
            SELECT
                p.name AS project,
                p.customer,
                COALESCE(budget.total_budget, 0) AS budget,
                CASE WHEN COALESCE(budget.total_budget, 0) > 0
                     THEN ROUND(COALESCE(fact.total_fact, 0) * 100.0
                          / budget.total_budget, 1)
                     ELSE 0
                END AS budget_pct,
                COALESCE(id_ready.pct, 0) AS id_readiness_pct,
                COALESCE(insp.open_count, 0) AS open_inspections,
                p.status,
                p.planned_end
            FROM projects p
            LEFT JOIN LATERAL (
                SELECT SUM(se.total_amount) AS total_budget
                FROM summary_estimates se WHERE se.project_id = p.id
            ) budget ON TRUE
            LEFT JOIN LATERAL (
                SELECT SUM(pr.amount_fact) AS total_fact
                FROM progress pr
                JOIN estimate_items ei ON ei.id = pr.estimate_item_id
                JOIN local_estimates le ON le.id = ei.ls_id
                JOIN local_estimate_bases leb ON leb.id = le.lsr_id
                JOIN object_estimates oe ON oe.id = leb.object_estimate_id
                JOIN summary_estimates se ON se.id = oe.summary_estimate_id
                WHERE se.project_id = p.id
            ) fact ON TRUE
            LEFT JOIN LATERAL (
                SELECT
                    CASE WHEN COUNT(d.id) > 0
                         THEN ROUND(COUNT(DISTINCT CASE WHEN ad.status = 'SIGNED'
                              THEN ad.id END) * 100.0 / COUNT(DISTINCT d.id), 1)
                         ELSE 0
                    END AS pct
                FROM drawings d
                LEFT JOIN asbuilt_drawing abd ON abd.drawing_id = d.id
                LEFT JOIN asbuilt_docs ad ON ad.id = abd.asbuilt_doc_id
                WHERE d.project_id = p.id
            ) id_ready ON TRUE
            LEFT JOIN LATERAL (
                SELECT COUNT(*) AS open_count
                FROM inspections i
                WHERE i.project_id = p.id AND i.status != 'CLOSED'
            ) insp ON TRUE
            ORDER BY p.name
        """, {})

        columns = [
            ("project", "Проект"), ("customer", "Заказчик"),
            ("budget", "Бюджет"), ("budget_pct", "Освоено (%)"),
            ("id_readiness_pct", "ИД (%)"),
            ("open_inspections", "Откр. предп."),
            ("status", "Статус"), ("planned_end", "Дата сдачи"),
        ]
        num_cols = {"budget", "budget_pct", "id_readiness_pct", "open_inspections"}

        self.add_header("Сводная по всем проектам")
        self.add_table(rows, columns, number_columns=num_cols)
        self.set_column_widths([30, 25, 18, 12, 10, 14, 14, 14])
        return self.save_to_bytes()


# ═══════════════════════════════════════════════════════════════
# Registry: ReportType → ExcelReport subclass
# ═══════════════════════════════════════════════════════════════

EXCEL_BUILDERS: dict[str, type[ExcelReport]] = {
    "BUDGET_BALANCE": BudgetBalanceReport,
    "BUDGET_BY_CONTRACTS": BudgetByContractsReport,
    "MONTHLY_PROGRESS": MonthlyProgressReport,
    "QUARTERLY_PROGRESS": QuarterlyProgressReport,
    "KS6A": KS6AReport,
    "M29_REPORT": M29ExcelReport,
    "MATERIAL_DEMAND": MaterialDemandReport,
    "ID_REGISTRY": IDRegistryReport,
    "ID_READINESS": IDReadinessReport,
    "INSPECTION_REPORT": InspectionExcelReport,
    "CERTIFICATE_STATUS": CertificateStatusReport,
    "SCHEDULE_DEVIATION": ScheduleDeviationReport,
    "INPUT_CONTROL_JOURNAL": InputControlJournalReport,
    "CONTRACT_REGISTRY": ContractRegistryReport,
    "MARGIN_REPORT": MarginReport,
    "UNCOVERED_LS": UncoveredLSReport,
    "MULTI_PROJECT_SUMMARY": MultiProjectSummaryReport,
}
